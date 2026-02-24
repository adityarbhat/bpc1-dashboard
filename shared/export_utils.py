"""
Centralized export utilities for creating multi-sheet Excel files
with all 7 group comparison pages.

This module is lazy-loaded only when the export page is accessed
to avoid performance impact on other pages.
"""

import pandas as pd
import io
from datetime import datetime


def format_period_for_airtable(period: str, year: int) -> str:
    """
    Convert session state period to Airtable format

    Args:
        period: Session state period ('year_end' or 'june_end')
        year: Year as integer (e.g., 2024)

    Returns:
        Formatted period string for Airtable
        - 'year_end' + 2024 → "2024 Annual"
        - 'june_end' + 2024 → "June 2024"
    """
    if period == 'year_end':
        return f"{year} Annual"
    else:  # june_end
        return f"June {year}"


def generate_filename(period: str, year: int) -> str:
    """
    Generate standardized filename for Excel export

    Args:
        period: Session state period ('year_end' or 'june_end')
        year: Year as integer (e.g., 2024)

    Returns:
        Formatted filename
        - Example: "BPC_Group_Analysis_2024_YearEnd.xlsx"
    """
    period_text = "YearEnd" if period == 'year_end' else "MidYear"
    timestamp = datetime.now().strftime("%Y%m%d")
    return f"BPC_Group_Analysis_{year}_{period_text}_{timestamp}.xlsx"


def create_multi_sheet_export(period: str, year: int) -> bytes:
    """
    Create multi-sheet Excel file with all 7 group comparison pages
    NOW WITH PROFESSIONAL FORMATTING

    Args:
        period: Session state period ('year_end' or 'june_end')
        year: Year as integer (e.g., 2024)

    Returns:
        Excel file as bytes for st.download_button with professional formatting

    Sheets created:
        1. Ratios - Color-coded performance indicators
        2. Balance Sheet - Winner highlighting
        3. Income Statement - Winner highlighting
        4. Labor Cost - Professional styling
        5. Business Mix - Professional styling
        6. Value - Winner highlighting
        7. Cash Flow - Professional styling

    Formatting includes:
        - Atlas branding (colors, fonts)
        - Color-coded performance thresholds (green/yellow/red)
        - Winner highlighting (green for best performers)
        - Metadata headers with period/year
        - Color legends
        - Frozen header rows
        - Auto-width columns
        - Professional borders and alignment
    """
    # Convert to Airtable format
    airtable_period = format_period_for_airtable(period, year)

    # Import extraction functions (lazy loading)
    from pages.group_pages.group_ratios import extract_ratios_data_for_export
    from pages.group_pages.group_balance_sheet import extract_balance_sheet_data_for_export
    from pages.group_pages.group_income_statement import extract_income_statement_data_for_export
    from pages.group_pages.group_labor_cost import extract_labor_cost_data_for_export
    from pages.group_pages.group_business_mix import extract_business_mix_data_for_export
    from pages.group_pages.group_value import extract_value_data_for_export
    from pages.group_pages.group_cash_flow import extract_cash_flow_data_for_export

    # Create in-memory Excel file
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:

        # Sheet 1: Ratios
        try:
            ratios_df = extract_ratios_data_for_export(airtable_period)
            if ratios_df is not None and not ratios_df.empty:
                ratios_df.to_excel(writer, sheet_name='Ratios', index=True)
        except Exception as e:
            print(f"Warning: Failed to export Ratios sheet: {e}")

        # Sheet 2: Balance Sheet
        try:
            balance_sheet_df = extract_balance_sheet_data_for_export(airtable_period)
            if balance_sheet_df is not None and not balance_sheet_df.empty:
                balance_sheet_df.to_excel(writer, sheet_name='Balance Sheet', index=True)
        except Exception as e:
            print(f"Warning: Failed to export Balance Sheet sheet: {e}")

        # Sheet 3: Income Statement
        try:
            income_statement_df = extract_income_statement_data_for_export(airtable_period)
            if income_statement_df is not None and not income_statement_df.empty:
                income_statement_df.to_excel(writer, sheet_name='Income Statement', index=True)
        except Exception as e:
            print(f"Warning: Failed to export Income Statement sheet: {e}")

        # Sheet 4: Labor Cost
        try:
            labor_cost_df = extract_labor_cost_data_for_export(airtable_period)
            if labor_cost_df is not None and not labor_cost_df.empty:
                labor_cost_df.to_excel(writer, sheet_name='Labor Cost', index=True)
        except Exception as e:
            print(f"Warning: Failed to export Labor Cost sheet: {e}")

        # Sheet 5: Business Mix
        try:
            business_mix_df = extract_business_mix_data_for_export(airtable_period)
            if business_mix_df is not None and not business_mix_df.empty:
                business_mix_df.to_excel(writer, sheet_name='Business Mix', index=True)
        except Exception as e:
            print(f"Warning: Failed to export Business Mix sheet: {e}")

        # Sheet 6: Value
        try:
            value_df = extract_value_data_for_export(airtable_period)
            if value_df is not None and not value_df.empty:
                value_df.to_excel(writer, sheet_name='Value', index=True)
        except Exception as e:
            print(f"Warning: Failed to export Value sheet: {e}")

        # Sheet 7: Cash Flow
        try:
            cash_flow_df = extract_cash_flow_data_for_export(airtable_period)
            if cash_flow_df is not None and not cash_flow_df.empty:
                cash_flow_df.to_excel(writer, sheet_name='Cash Flow', index=True)
        except Exception as e:
            print(f"Warning: Failed to export Cash Flow sheet: {e}")

    # ===== FORMATTING PHASE =====
    # Re-open workbook and apply professional formatting to all sheets

    output.seek(0)
    from openpyxl import load_workbook

    # Import formatting functions (lazy loading)
    from shared.excel_formatter import (
        format_ratios_sheet,
        format_balance_sheet_sheet,
        format_income_statement_sheet,
        format_labor_cost_sheet,
        format_business_mix_sheet,
        format_value_sheet,
        format_cash_flow_sheet
    )

    try:
        wb = load_workbook(output)

        # Format each sheet with error handling
        try:
            if 'Ratios' in wb.sheetnames:
                format_ratios_sheet(wb, 'Ratios', period, year)
        except Exception as e:
            print(f"Warning: Failed to format Ratios sheet: {e}")

        try:
            if 'Balance Sheet' in wb.sheetnames:
                format_balance_sheet_sheet(wb, 'Balance Sheet', period, year)
        except Exception as e:
            print(f"Warning: Failed to format Balance Sheet sheet: {e}")

        try:
            if 'Income Statement' in wb.sheetnames:
                format_income_statement_sheet(wb, 'Income Statement', period, year)
        except Exception as e:
            print(f"Warning: Failed to format Income Statement sheet: {e}")

        try:
            if 'Labor Cost' in wb.sheetnames:
                format_labor_cost_sheet(wb, 'Labor Cost', period, year)
        except Exception as e:
            print(f"Warning: Failed to format Labor Cost sheet: {e}")

        try:
            if 'Business Mix' in wb.sheetnames:
                format_business_mix_sheet(wb, 'Business Mix', period, year)
        except Exception as e:
            print(f"Warning: Failed to format Business Mix sheet: {e}")

        try:
            if 'Value' in wb.sheetnames:
                format_value_sheet(wb, 'Value', period, year)
        except Exception as e:
            print(f"Warning: Failed to format Value sheet: {e}")

        try:
            if 'Cash Flow' in wb.sheetnames:
                format_cash_flow_sheet(wb, 'Cash Flow', period, year)
        except Exception as e:
            print(f"Warning: Failed to format Cash Flow sheet: {e}")

        # Save formatted workbook back to bytes
        formatted_output = io.BytesIO()
        wb.save(formatted_output)
        formatted_output.seek(0)
        return formatted_output.getvalue()

    except Exception as e:
        # If formatting fails completely, return unformatted workbook
        print(f"Warning: Excel formatting failed: {e}")
        print("Returning unformatted Excel file...")
        output.seek(0)
        return output.getvalue()
