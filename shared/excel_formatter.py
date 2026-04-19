"""
Professional Excel formatting for BPC Dashboard exports.
Applies Atlas branding, color coding, and professional styling to all 7 sheets.

This module is lazy-loaded only when the export page is accessed
to avoid performance impact on other pages.
"""

from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime


# ===== ATLAS BRANDING CONSTANTS =====
ATLAS_BLUE = "025a9a"
ATLAS_LIGHT_BLUE = "0e9cd5"
ATLAS_RED = "c2002f"
GREEN_GREAT = "c8e6c9"         # Matches web: #c8e6c9
YELLOW_CAUTION = "fff3c4"      # Matches web: #fff3c4
RED_IMPROVE = "ffcdd2"         # Matches web: #ffcdd2
DARK_GREY = "4a5568"           # Category headers
LIGHT_GREY = "f7fafc"          # Row label background
WINNER_GREEN = "c6f6d5"        # Winner cells (Balance Sheet, Value)
ABOVE_AVG_YELLOW = "ffe082"    # Above average cells


# ===== FONT DEFINITIONS =====
HEADER_FONT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
TITLE_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
COLUMN_HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
CATEGORY_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Calibri', size=10)
BOLD_DATA_FONT = Font(name='Calibri', size=10, bold=True)


# ===== BORDER DEFINITIONS =====
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


# ===== RATIO COLOR THRESHOLDS =====
# Matches group_ratios.py get_cell_color() function (lines 307-319)
RATIO_THRESHOLDS = {
    'Current Ratio': {'great': 2.0, 'caution': [1.2, 2.0], 'reverse': False},
    'Debt to Equity': {'great': 1.4, 'caution': [1.5, 2.9], 'reverse': True},
    'Working Capital %': {'great': 0.30, 'caution': [0.15, 0.2999], 'reverse': False},
    'Survival Score': {'great': 3.0, 'caution': [2.0, 3.0], 'reverse': False},
    'Sales/Assets': {'great': 3.7, 'caution': [2.0, 3.6], 'reverse': False},
    'Gross Profit Margin': {'great': 0.25, 'caution': [0.20, 0.25], 'reverse': False},
    'Operating Profit Margin': {'great': 0.055, 'caution': [0.03, 0.0549], 'reverse': False},
    'Net Profit Margin': {'great': 0.05, 'caution': [0.03, 0.0499], 'reverse': False},
    'Revenue Per Employee': {'great': 550, 'caution': [325, 550], 'reverse': False},
    'EBITDA/Revenue': {'great': 0.05, 'caution': [0.025, 0.05], 'reverse': False},
    'Days Sales Outstanding (DSO)': {'great': 30, 'caution': [30, 60], 'reverse': True},
    'OCF/Revenue': {'great': 0.0, 'caution': [-0.03, 0.0], 'reverse': False},
    'FCF/Revenue': {'great': 0.005, 'caution': [-0.005, 0.005], 'reverse': False},
    'NCF/Revenue': {'great': 0.005, 'caution': [-0.005, 0.005], 'reverse': False},
}


# ===== HELPER FUNCTIONS =====

def get_ratio_color(value, metric_name):
    """
    Get fill color for ratio cells based on performance thresholds.
    Returns openpyxl PatternFill object.

    Args:
        value: Numeric value or string representation
        metric_name: Name of the metric (e.g., "Current Ratio")

    Returns:
        PatternFill object with appropriate color
    """
    # Handle missing or zero values
    if value is None or value == '' or value == 0 or value == '-':
        return PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")

    # Unknown metric
    if metric_name not in RATIO_THRESHOLDS:
        return None  # No fill for unknown metrics

    threshold = RATIO_THRESHOLDS[metric_name]
    is_reverse = threshold.get('reverse', False)

    try:
        # Parse value - handle formatted strings like "2.5", "35.2%", "$580K"
        val_str = str(value).strip()

        # Remove percentage sign
        if '%' in val_str:
            val_str = val_str.replace('%', '')
            val = float(val_str) / 100  # Convert to decimal
        # Remove dollar sign and K suffix
        elif '$' in val_str and 'K' in val_str:
            val_str = val_str.replace('$', '').replace('K', '').replace(',', '')
            val = float(val_str)  # Already in thousands
        # Remove dollar sign and M suffix
        elif '$' in val_str and 'M' in val_str:
            val_str = val_str.replace('$', '').replace('M', '').replace(',', '')
            val = float(val_str) * 1000  # Convert to thousands
        # Remove dollar sign and commas
        elif '$' in val_str:
            val_str = val_str.replace('$', '').replace(',', '')
            val = float(val_str) / 1000  # Convert to thousands
        # Plain number
        else:
            val_str = val_str.replace(',', '')
            val = float(val_str)

        # Apply threshold logic
        if is_reverse:
            # Lower is better
            if val <= threshold['great']:
                return PatternFill(start_color=GREEN_GREAT, end_color=GREEN_GREAT, fill_type="solid")
            elif isinstance(threshold['caution'], list) and threshold['caution'][0] <= val <= threshold['caution'][1]:
                return PatternFill(start_color=YELLOW_CAUTION, end_color=YELLOW_CAUTION, fill_type="solid")
            else:
                return PatternFill(start_color=RED_IMPROVE, end_color=RED_IMPROVE, fill_type="solid")
        else:
            # Higher is better
            if val >= threshold['great']:
                return PatternFill(start_color=GREEN_GREAT, end_color=GREEN_GREAT, fill_type="solid")
            elif isinstance(threshold['caution'], list) and threshold['caution'][0] <= val <= threshold['caution'][1]:
                return PatternFill(start_color=YELLOW_CAUTION, end_color=YELLOW_CAUTION, fill_type="solid")
            else:
                return PatternFill(start_color=RED_IMPROVE, end_color=RED_IMPROVE, fill_type="solid")

    except (ValueError, TypeError) as e:
        # Return light gray for unparseable values
        return PatternFill(start_color="f8f9fa", end_color="f8f9fa", fill_type="solid")


def add_metadata_header(ws, sheet_name, period, year):
    """
    Add title row with period/year information at the top of sheet.
    Example: "BPC Group Analysis - Ratios - 2024 Year End"

    Args:
        ws: openpyxl worksheet object
        sheet_name: Name of the sheet (e.g., "Ratios")
        period: Session state period ('year_end' or 'june_end')
        year: Year as integer (e.g., 2024)
    """
    # Insert a row at the top
    ws.insert_rows(1)

    # Format period text
    period_text = "Year End" if period == 'year_end' else "Mid Year"

    # Create title
    title = f"BPC Group Analysis - {sheet_name} - {year} {period_text}"

    # Set value in A1
    ws['A1'] = title

    # Merge across all columns
    max_col = ws.max_column
    ws.merge_cells(f'A1:{get_column_letter(max_col)}1')

    # Apply formatting
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = PatternFill(start_color=ATLAS_BLUE, end_color=ATLAS_BLUE, fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25


def add_color_legend(ws, sheet_type):
    """
    Add color legend explaining what colors mean.
    Different legends for different sheet types.

    Args:
        ws: openpyxl worksheet object
        sheet_type: Type of sheet ('ratios', 'balance_sheet', 'income_statement', 'value', 'labor_cost', 'cash_flow', 'other')
    """
    # Find the last row with data
    last_row = ws.max_row

    # Add 2 blank rows
    legend_row = last_row + 3

    # Legend title
    ws[f'A{legend_row}'] = "Color Legend:"
    ws[f'A{legend_row}'].font = BOLD_DATA_FONT

    legend_row += 1

    if sheet_type == 'ratios':
        # Green = Great, Yellow = Caution, Red = Improve
        ws[f'A{legend_row}'] = "Great Performance"
        ws[f'A{legend_row}'].fill = PatternFill(start_color=GREEN_GREAT, end_color=GREEN_GREAT, fill_type="solid")
        ws[f'A{legend_row}'].font = DATA_FONT
        ws[f'A{legend_row}'].border = THIN_BORDER

        ws[f'B{legend_row}'] = "Caution"
        ws[f'B{legend_row}'].fill = PatternFill(start_color=YELLOW_CAUTION, end_color=YELLOW_CAUTION, fill_type="solid")
        ws[f'B{legend_row}'].font = DATA_FONT
        ws[f'B{legend_row}'].border = THIN_BORDER

        ws[f'C{legend_row}'] = "Needs Improvement"
        ws[f'C{legend_row}'].fill = PatternFill(start_color=RED_IMPROVE, end_color=RED_IMPROVE, fill_type="solid")
        ws[f'C{legend_row}'].font = DATA_FONT
        ws[f'C{legend_row}'].border = THIN_BORDER

    elif sheet_type == 'balance_sheet':
        # Green = Winner, Yellow = Above Average (for liabilities)
        ws[f'A{legend_row}'] = "Leading Agent (Best Performance)"
        ws[f'A{legend_row}'].fill = PatternFill(start_color=WINNER_GREEN, end_color=WINNER_GREEN, fill_type="solid")
        ws[f'A{legend_row}'].font = DATA_FONT
        ws[f'A{legend_row}'].border = THIN_BORDER

        ws[f'B{legend_row}'] = "Above Group Average (Liabilities)"
        ws[f'B{legend_row}'].fill = PatternFill(start_color=ABOVE_AVG_YELLOW, end_color=ABOVE_AVG_YELLOW, fill_type="solid")
        ws[f'B{legend_row}'].font = DATA_FONT
        ws[f'B{legend_row}'].border = THIN_BORDER

    elif sheet_type == 'income_statement':
        # Green = Winner, Yellow = Above Average (for expenses)
        ws[f'A{legend_row}'] = "Leading Agent (Highest Revenue)"
        ws[f'A{legend_row}'].fill = PatternFill(start_color=WINNER_GREEN, end_color=WINNER_GREEN, fill_type="solid")
        ws[f'A{legend_row}'].font = DATA_FONT
        ws[f'A{legend_row}'].border = THIN_BORDER

        ws[f'B{legend_row}'] = "Above Group Average (Expenses)"
        ws[f'B{legend_row}'].fill = PatternFill(start_color=ABOVE_AVG_YELLOW, end_color=ABOVE_AVG_YELLOW, fill_type="solid")
        ws[f'B{legend_row}'].font = DATA_FONT
        ws[f'B{legend_row}'].border = THIN_BORDER

    elif sheet_type == 'labor_cost':
        # Green = Winner (lowest cost), Yellow = Above Average
        ws[f'A{legend_row}'] = "Leading Agent (Lowest Cost)"
        ws[f'A{legend_row}'].fill = PatternFill(start_color=WINNER_GREEN, end_color=WINNER_GREEN, fill_type="solid")
        ws[f'A{legend_row}'].font = DATA_FONT
        ws[f'A{legend_row}'].border = THIN_BORDER

        ws[f'B{legend_row}'] = "Above Group Average"
        ws[f'B{legend_row}'].fill = PatternFill(start_color=ABOVE_AVG_YELLOW, end_color=ABOVE_AVG_YELLOW, fill_type="solid")
        ws[f'B{legend_row}'].font = DATA_FONT
        ws[f'B{legend_row}'].border = THIN_BORDER

    elif sheet_type == 'cash_flow':
        # Green = Winner (highest cash flow), Yellow = Above Average
        ws[f'A{legend_row}'] = "Leading Agent (Best Cash Flow)"
        ws[f'A{legend_row}'].fill = PatternFill(start_color=WINNER_GREEN, end_color=WINNER_GREEN, fill_type="solid")
        ws[f'A{legend_row}'].font = DATA_FONT
        ws[f'A{legend_row}'].border = THIN_BORDER

        ws[f'B{legend_row}'] = "Above Group Average"
        ws[f'B{legend_row}'].fill = PatternFill(start_color=ABOVE_AVG_YELLOW, end_color=ABOVE_AVG_YELLOW, fill_type="solid")
        ws[f'B{legend_row}'].font = DATA_FONT
        ws[f'B{legend_row}'].border = THIN_BORDER

    elif sheet_type == 'value':
        # Green = Winner only
        ws[f'A{legend_row}'] = "Winner (Best Performance)"
        ws[f'A{legend_row}'].fill = PatternFill(start_color=WINNER_GREEN, end_color=WINNER_GREEN, fill_type="solid")
        ws[f'A{legend_row}'].font = DATA_FONT
        ws[f'A{legend_row}'].border = THIN_BORDER


def freeze_top_rows(ws, num_rows=3):
    """
    Freeze top N rows for easy scrolling.

    Args:
        ws: openpyxl worksheet object
        num_rows: Number of rows to freeze (default 3: title + blank + headers)
    """
    ws.freeze_panes = f'A{num_rows + 1}'


def auto_width_columns(ws, min_width=12, max_width=25):
    """
    Set column widths based on content with min/max constraints.

    Args:
        ws: openpyxl worksheet object
        min_width: Minimum column width (default 12)
        max_width: Maximum column width (default 25)
    """
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)

        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass

        # Add padding and apply min/max
        adjusted_width = min(max(max_length + 2, min_width), max_width)
        ws.column_dimensions[column_letter].width = adjusted_width


def apply_borders_to_range(ws, start_row, end_row, start_col, end_col):
    """
    Apply thin borders to a range of cells.

    Args:
        ws: openpyxl worksheet object
        start_row: Starting row number
        end_row: Ending row number
        start_col: Starting column number
        end_col: Ending column number
    """
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


# ===== SHEET-SPECIFIC FORMATTING FUNCTIONS =====

def format_ratios_sheet(workbook, sheet_name, period, year):
    """
    Format the Ratios sheet with color-coded performance indicators.

    Sheet structure after pandas.to_excel():
    - Row 1: Column headers (index name, company names)
    - Rows 2+: Data rows with ratio names in column A

    Formatting applied:
    1. Insert metadata header at row 1
    2. Apply Atlas blue to column headers
    3. Color-code data cells based on RATIO_THRESHOLDS
    4. Bold row labels (column A)
    5. Freeze top 3 rows
    6. Auto-width columns
    7. Add color legend at bottom
    """
    ws = workbook[sheet_name]

    # Step 1: Add metadata header (inserts new row 1)
    add_metadata_header(ws, sheet_name, period, year)

    # After inserting header:
    # Row 1: Title
    # Row 2: Column headers (was row 1)
    # Row 3+: Data rows (was row 2+)

    # Step 2: Format column headers (now row 2)
    header_row = 2
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = COLUMN_HEADER_FONT
        cell.fill = PatternFill(start_color=ATLAS_BLUE, end_color=ATLAS_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    ws.row_dimensions[header_row].height = 20

    # Step 3 & 4: Format data rows (row 3 onwards)
    for row in range(3, ws.max_row + 1):
        # Column A: Row label (metric name)
        metric_name_cell = ws.cell(row=row, column=1)
        metric_name = metric_name_cell.value

        if metric_name:
            # Bold the metric name
            metric_name_cell.font = BOLD_DATA_FONT
            metric_name_cell.alignment = Alignment(horizontal='left', vertical='center')
            metric_name_cell.border = THIN_BORDER

            # Color-code data cells (columns 2 onwards)
            for col in range(2, ws.max_column + 1):
                data_cell = ws.cell(row=row, column=col)
                value = data_cell.value

                # Apply color based on threshold
                fill = get_ratio_color(value, metric_name)
                if fill:
                    data_cell.fill = fill

                data_cell.font = DATA_FONT
                data_cell.alignment = Alignment(horizontal='center', vertical='center')
                data_cell.border = THIN_BORDER

    # Step 5: Freeze top 3 rows (title + blank space + headers)
    freeze_top_rows(ws, num_rows=2)

    # Step 6: Auto-width columns
    auto_width_columns(ws, min_width=15, max_width=20)

    # Step 7: Add color legend
    add_color_legend(ws, 'ratios')


def format_balance_sheet_sheet(workbook, sheet_name, period, year):
    """
    Format Balance Sheet sheet with winner highlighting and above-average yellow.

    Formatting logic:
    - Green: Winner cells (highest % for assets, lowest % for liabilities)
    - Yellow: Liability items above group average
    - Atlas blue column headers
    """
    ws = workbook[sheet_name]

    # Step 1: Add metadata header
    add_metadata_header(ws, sheet_name, period, year)

    # Step 2: Format column headers (now row 2)
    header_row = 2
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = COLUMN_HEADER_FONT
        cell.fill = PatternFill(start_color=ATLAS_BLUE, end_color=ATLAS_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    ws.row_dimensions[header_row].height = 20

    # Define items where HIGHER is better (highlight highest in green)
    high_is_good_items = [
        'Cash and Cash Equivalents',
        'Total Current Assets',
        'Gross Fixed Assets',
        'Net Fixed Assets',
        "Owners' Equity",
        "Owner's Equity"
    ]

    # Define informational items (highlight highest in green)
    informational_items = [
        'Trade Accounts Receivable',
        'Receivables',
        'Other Receivables',
        'Prepaid Expenses',
        'Related Company Receivables',
        'Owner Receivables',
        'Other Current Assets',
        'Inter Company Receivables',
        'Other Assets',
    ]

    # Define items where LOWER is better (Accumulated Depreciation)
    low_is_good_asset_items = [
        'Accumulated Depreciation (-)',
    ]

    # Define liability items where LOWER is better and above-average gets yellow
    liability_items = [
        'Notes Payable - Bank',
        'Notes Payable to Owners - Current Portion',
        'Trade Accounts Payable',
        'Accrued Expenses',
        'Current Portion LTD',
        'Inter Company Payable',
        'Other Current Liabilities',
        'Total Current Liabilities',
        'Long-Term Debt',
        'Notes Payable to Owners - LT',
        'Inter Company Debt',
        'Other LT Liabilities',
        'Long Term Liabilities',
        'Total Liabilities'
    ]

    # Step 3: First pass - calculate group averages for each row
    row_averages = {}
    for row in range(3, ws.max_row + 1):
        row_label = ws.cell(row=row, column=1).value
        if row_label:
            row_values = []
            for col in range(2, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and cell_value != '-':
                    try:
                        val_str = str(cell_value).strip().replace('%', '').replace('$', '').replace(',', '')
                        if '(' in val_str:
                            val_str = val_str.replace('(', '-').replace(')', '')
                        row_values.append(float(val_str))
                    except:
                        pass
            if row_values:
                row_averages[row] = sum(row_values) / len(row_values)

    # Step 4: Format data rows with winner and above-average highlighting
    for row in range(3, ws.max_row + 1):
        row_label = ws.cell(row=row, column=1).value

        if row_label:
            # Bold the row label
            ws.cell(row=row, column=1).font = BOLD_DATA_FONT
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row, column=1).border = THIN_BORDER

            # Collect row values for winner detection
            row_values = []
            for col in range(2, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and cell_value != '-':
                    try:
                        val_str = str(cell_value).strip().replace('%', '').replace('$', '').replace(',', '')
                        if '(' in val_str:
                            val_str = val_str.replace('(', '-').replace(')', '')
                        row_values.append(float(val_str))
                    except:
                        row_values.append(0)
                else:
                    row_values.append(0)

            # Determine winner based on item type
            winner_value = None
            is_liability = any(str(row_label).strip() == item or str(row_label).strip().startswith(item) for item in liability_items)
            is_high_is_good = any(str(row_label).strip() == item or str(row_label).strip().startswith(item) for item in high_is_good_items + informational_items)
            is_low_is_good_asset = any(str(row_label).strip() == item or str(row_label).strip().startswith(item) for item in low_is_good_asset_items)

            if row_values:
                non_zero_values = [v for v in row_values if v != 0]
                if non_zero_values:
                    if is_liability or is_low_is_good_asset:
                        winner_value = min(non_zero_values)
                    elif is_high_is_good:
                        winner_value = max(non_zero_values)
                    else:
                        winner_value = max(non_zero_values)

            # Get group average for this row
            group_avg = row_averages.get(row, 0)

            # Apply formatting to data cells
            for col_idx, col in enumerate(range(2, ws.max_column + 1)):
                data_cell = ws.cell(row=row, column=col)
                cell_value = data_cell.value

                # Parse value
                parsed_value = None
                try:
                    val_str = str(cell_value).strip().replace('%', '').replace('$', '').replace(',', '')
                    if '(' in val_str:
                        val_str = val_str.replace('(', '-').replace(')', '')
                    parsed_value = float(val_str)
                except:
                    pass

                # Apply highlighting
                if parsed_value is not None and parsed_value != 0:
                    # Check if winner
                    if winner_value is not None and parsed_value == winner_value:
                        data_cell.fill = PatternFill(start_color=WINNER_GREEN, end_color=WINNER_GREEN, fill_type="solid")
                    # Check if liability and above average (yellow)
                    elif is_liability and parsed_value > group_avg:
                        data_cell.fill = PatternFill(start_color=ABOVE_AVG_YELLOW, end_color=ABOVE_AVG_YELLOW, fill_type="solid")

                data_cell.font = DATA_FONT
                data_cell.alignment = Alignment(horizontal='center', vertical='center')
                data_cell.border = THIN_BORDER

    # Freeze and auto-width
    freeze_top_rows(ws, num_rows=2)
    auto_width_columns(ws, min_width=15, max_width=25)

    # Add legend
    add_color_legend(ws, 'balance_sheet')


def format_income_statement_sheet(workbook, sheet_name, period, year):
    """
    Format Income Statement sheet with winner and above-average highlighting.
    - Green: Winner (highest for revenue items)
    - Yellow: Above group average for expense items
    """
    ws = workbook[sheet_name]

    # Add metadata header
    add_metadata_header(ws, sheet_name, period, year)

    # Format column headers (now row 2)
    header_row = 2
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = COLUMN_HEADER_FONT
        cell.fill = PatternFill(start_color=ATLAS_BLUE, end_color=ATLAS_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    ws.row_dimensions[header_row].height = 20

    # Define revenue line items (highest is best - green winner)
    revenue_line_items = [
        'Intra State HHG', 'Local HHG', 'Inter State HHG', 'Office & Industrial',
        'Warehouse (Non-commercial)', 'Warehouse handling (Non-commercial)', 'International',
        'Packing & Unpacking', 'Booking & Royalties', 'Special Products', 'Records Storage',
        'Military DPM Contracts', 'Distribution', 'Hotel Deliveries', 'Other Revenue'
    ]

    # Define other income items (highest is best - green winner)
    other_income_line_items = [
        'PPP Funds Received (forgiven)', 'Other Income'
    ]

    # Define expense items (above average gets yellow)
    expense_line_items = [
        'Direct Wages', 'Vehicle Operating Expense', 'Packing & Warehouse Supplies',
        'Owner Operator - Intra State', 'Owner Operator - Inter State',
        'Owner Operator - Office & Industrial', 'Owner Operator - Packing',
        'Owner Operator - General & Other', 'Claims', 'Other Transportation Expense',
        'Lease Expense - Revenue Equipment', 'Other Direct Expenses',
        'Rent and/or Building Expense', 'Depreciation/Amortization', 'Total Cost Of Revenue',
        'Advertising & Marketing', 'Bad Debts', 'Sales Compensation', 'Contributions',
        'Computer Support', 'Dues & Subscriptions', 'Payroll Taxes & Benefits',
        'Lease Expense - Office Equipment', "Workers' Comp. Insurance", 'Insurance',
        'Legal & Accounting', 'Office Expense', 'Other Administrative',
        'Pension, profit sharing, 401K', 'Professional Fees', 'Repairs & Maintenance',
        'Salaries - Administrative', 'Taxes & Licenses', 'Telephone/Fax/Utilities/Internet',
        'Travel & Entertainment', 'Vehicle Expense - Administrative', 'Total Operating Expenses',
        'CEO Comp/Perks (-)', 'Other Expense (-)', 'Interest Expense (-)'
    ]

    # First pass - calculate group averages for each row
    row_averages = {}
    for row in range(3, ws.max_row + 1):
        row_label = ws.cell(row=row, column=1).value
        if row_label:
            row_values = []
            for col in range(2, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and cell_value != '-':
                    try:
                        val_str = str(cell_value).strip().replace('%', '').replace('$', '').replace(',', '')
                        if '(' in val_str:
                            val_str = val_str.replace('(', '-').replace(')', '')
                        row_values.append(float(val_str))
                    except:
                        pass
            if row_values:
                row_averages[row] = sum(row_values) / len(row_values)

    # Format data rows with winner and above-average highlighting
    for row in range(3, ws.max_row + 1):
        row_label = ws.cell(row=row, column=1).value

        if row_label:
            # Bold the row label
            ws.cell(row=row, column=1).font = BOLD_DATA_FONT
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row, column=1).border = THIN_BORDER

            # Collect row values
            row_values = []
            for col in range(2, ws.max_column + 1):
                cell_value = ws.cell(row=row, column=col).value
                if cell_value and cell_value != '-':
                    try:
                        val_str = str(cell_value).strip().replace('%', '').replace('$', '').replace(',', '')
                        if '(' in val_str:
                            val_str = val_str.replace('(', '-').replace(')', '')
                        row_values.append(float(val_str))
                    except:
                        row_values.append(0)
                else:
                    row_values.append(0)

            # Check item type
            is_revenue = any(str(row_label).strip() == item or str(row_label).strip().startswith(item) for item in revenue_line_items)
            is_other_income = any(str(row_label).strip() == item or str(row_label).strip().startswith(item) for item in other_income_line_items)
            is_expense = any(str(row_label).strip() == item or str(row_label).strip().startswith(item) for item in expense_line_items)

            # Find winner value (max for revenue/income items)
            winner_value = None
            if row_values:
                non_zero_values = [v for v in row_values if v != 0]
                if non_zero_values and (is_revenue or is_other_income):
                    winner_value = max(non_zero_values)

            # Get group average
            group_avg = row_averages.get(row, 0)

            # Apply formatting
            for col_idx, col in enumerate(range(2, ws.max_column + 1)):
                data_cell = ws.cell(row=row, column=col)
                cell_value = data_cell.value

                parsed_value = None
                try:
                    val_str = str(cell_value).strip().replace('%', '').replace('$', '').replace(',', '')
                    if '(' in val_str:
                        val_str = val_str.replace('(', '-').replace(')', '')
                    parsed_value = float(val_str)
                except:
                    pass

                if parsed_value is not None and parsed_value != 0:
                    # Green for winner (revenue/income items)
                    if winner_value is not None and parsed_value == winner_value:
                        data_cell.fill = PatternFill(start_color=WINNER_GREEN, end_color=WINNER_GREEN, fill_type="solid")
                    # Yellow for expense items above average
                    elif is_expense:
                        # Handle negative expense items (CEO Comp, Other Expense, Interest Expense)
                        if any(neg_item in str(row_label) for neg_item in ['(-)']):
                            # More negative = worse, so compare < group_avg
                            if parsed_value < group_avg:
                                data_cell.fill = PatternFill(start_color=ABOVE_AVG_YELLOW, end_color=ABOVE_AVG_YELLOW, fill_type="solid")
                        else:
                            # Higher positive expense = worse
                            if parsed_value > group_avg:
                                data_cell.fill = PatternFill(start_color=ABOVE_AVG_YELLOW, end_color=ABOVE_AVG_YELLOW, fill_type="solid")

                data_cell.font = DATA_FONT
                data_cell.alignment = Alignment(horizontal='center', vertical='center')
                data_cell.border = THIN_BORDER

    freeze_top_rows(ws, num_rows=2)
    auto_width_columns(ws, min_width=15, max_width=25)
    add_color_legend(ws, 'income_statement')


def format_labor_cost_sheet(workbook, sheet_name, period, year):
    """
    Format Labor Cost sheet with winner and above-average highlighting.
    - Green: Winner (lowest percentage/cost = best)
    - Yellow: Above group average (higher cost = worse)
    """
    ws = workbook[sheet_name]

    # Add metadata header
    add_metadata_header(ws, sheet_name, period, year)

    # Format column headers (now row 2)
    header_row = 2
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = COLUMN_HEADER_FONT
        cell.fill = PatternFill(start_color=ATLAS_BLUE, end_color=ATLAS_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    ws.row_dimensions[header_row].height = 20

    # Identify rows where lower is better (matching web logic)
    percentage_row_keywords = ['% of Revenue', 'Labor Ratio']
    colorable_currency_items = ['Revenue Producing Labor and Expenses', 'Total Labor and Expenses']

    # First pass - calculate group averages for colorable rows
    row_averages = {}
    for row in range(3, ws.max_row + 1):
        row_label = ws.cell(row=row, column=1).value
        if row_label:
            is_percentage_row = any(keyword in str(row_label) for keyword in percentage_row_keywords)
            is_colorable_currency = any(str(row_label).startswith(item) for item in colorable_currency_items)

            if is_percentage_row or is_colorable_currency:
                row_values = []
                for col in range(2, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and cell_value != '-':
                        try:
                            val_str = str(cell_value).strip().replace('%', '').replace('$', '').replace(',', '')
                            if '(' in val_str:
                                val_str = val_str.replace('(', '-').replace(')', '')
                            row_values.append(float(val_str))
                        except:
                            pass
                if row_values:
                    row_averages[row] = sum(row_values) / len(row_values)

    # Format data rows with winner and above-average highlighting
    for row in range(3, ws.max_row + 1):
        row_label = ws.cell(row=row, column=1).value

        if row_label:
            # Bold the row label
            ws.cell(row=row, column=1).font = BOLD_DATA_FONT
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row, column=1).border = THIN_BORDER

            # Check if this row should be color-coded (lower is better)
            is_percentage_row = any(keyword in str(row_label) for keyword in percentage_row_keywords)
            is_colorable_currency = any(str(row_label).startswith(item) for item in colorable_currency_items)

            if is_percentage_row or is_colorable_currency:
                # Find MIN value in the row (excluding column A)
                row_values = []
                for col in range(2, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and cell_value != '-':
                        try:
                            val_str = str(cell_value).strip().replace('%', '').replace('$', '').replace(',', '')
                            if '(' in val_str:
                                val_str = val_str.replace('(', '-').replace(')', '')
                            row_values.append(float(val_str))
                        except:
                            row_values.append(float('inf'))
                    else:
                        row_values.append(float('inf'))

                # Find winner (minimum value = best)
                min_value = min(row_values) if row_values else float('inf')
                group_avg = row_averages.get(row, 0)

                for col_idx, col in enumerate(range(2, ws.max_column + 1)):
                    data_cell = ws.cell(row=row, column=col)
                    cell_value = data_cell.value

                    parsed_value = None
                    try:
                        val_str = str(cell_value).strip().replace('%', '').replace('$', '').replace(',', '')
                        if '(' in val_str:
                            val_str = val_str.replace('(', '-').replace(')', '')
                        parsed_value = float(val_str)
                    except:
                        pass

                    if parsed_value is not None and parsed_value != float('inf'):
                        # Green for winner (lowest)
                        if parsed_value == min_value and min_value != float('inf'):
                            data_cell.fill = PatternFill(start_color=WINNER_GREEN, end_color=WINNER_GREEN, fill_type="solid")
                        # Yellow for above average (worse)
                        elif parsed_value > group_avg:
                            data_cell.fill = PatternFill(start_color=ABOVE_AVG_YELLOW, end_color=ABOVE_AVG_YELLOW, fill_type="solid")

                    data_cell.font = DATA_FONT
                    data_cell.alignment = Alignment(horizontal='center', vertical='center')
                    data_cell.border = THIN_BORDER
            else:
                # Currency rows - just format, no highlighting
                for col in range(2, ws.max_column + 1):
                    data_cell = ws.cell(row=row, column=col)
                    data_cell.font = DATA_FONT
                    data_cell.alignment = Alignment(horizontal='center', vertical='center')
                    data_cell.border = THIN_BORDER

    freeze_top_rows(ws, num_rows=2)
    auto_width_columns(ws, min_width=15, max_width=25)
    add_color_legend(ws, 'labor_cost')


def format_business_mix_sheet(workbook, sheet_name, period, year):
    """
    Format Business Mix sheet.
    Percentage formatting with professional styling.
    """
    ws = workbook[sheet_name]

    # Add metadata header
    add_metadata_header(ws, sheet_name, period, year)

    # Format column headers (now row 2)
    header_row = 2
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = COLUMN_HEADER_FONT
        cell.fill = PatternFill(start_color=ATLAS_BLUE, end_color=ATLAS_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    ws.row_dimensions[header_row].height = 20

    # Format data rows
    for row in range(3, ws.max_row + 1):
        # Bold row labels
        ws.cell(row=row, column=1).font = BOLD_DATA_FONT
        ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
        ws.cell(row=row, column=1).border = THIN_BORDER

        # Format data cells
        for col in range(2, ws.max_column + 1):
            data_cell = ws.cell(row=row, column=col)
            data_cell.font = DATA_FONT
            data_cell.alignment = Alignment(horizontal='center', vertical='center')
            data_cell.border = THIN_BORDER

    freeze_top_rows(ws, num_rows=2)
    auto_width_columns(ws, min_width=15, max_width=20)


def format_value_sheet(workbook, sheet_name, period, year):
    """
    Format Value sheet with winner highlighting.
    For most items, HIGHEST value is best (green winner).
    For Interest Bearing Debt, LOWEST is best (green winner).
    Matches web page logic from group_value.py lines 303-306.
    """
    ws = workbook[sheet_name]

    # Add metadata header
    add_metadata_header(ws, sheet_name, period, year)

    # Format column headers (now row 2)
    header_row = 2
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = COLUMN_HEADER_FONT
        cell.fill = PatternFill(start_color=ATLAS_BLUE, end_color=ATLAS_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    ws.row_dimensions[header_row].height = 20

    # Define items where higher is better (matching web logic)
    high_is_good_items = ['EBITDA (000)', '3 x EBITDA (000)', 'Company Value (000)', 'Equity (000)', 'Value to Equity']

    # Define items where lower is better (matching web logic)
    low_is_good_items = ['Interest Bearing Debt (000)']

    # Format data rows with winner highlighting
    for row in range(3, ws.max_row + 1):
        row_label = ws.cell(row=row, column=1).value

        if row_label:
            # Bold the row label
            ws.cell(row=row, column=1).font = BOLD_DATA_FONT
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row, column=1).border = THIN_BORDER

            # Check which type of row this is
            is_high_is_good = any(str(row_label).startswith(item) or str(row_label) == item for item in high_is_good_items)
            is_low_is_good = any(str(row_label).startswith(item) or str(row_label) == item for item in low_is_good_items)

            if is_high_is_good or is_low_is_good:
                # Collect row values
                row_values = []
                for col in range(2, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and cell_value != '-':
                        try:
                            val_str = str(cell_value).strip().replace('$', '').replace(',', '')
                            if '(' in val_str:
                                val_str = val_str.replace('(', '-').replace(')', '')
                            row_values.append(float(val_str))
                        except:
                            row_values.append(float('-inf') if is_high_is_good else float('inf'))
                    else:
                        row_values.append(float('-inf') if is_high_is_good else float('inf'))

                # Find winner value
                if row_values:
                    if is_high_is_good:
                        winner_value = max(row_values)
                    else:  # is_low_is_good
                        winner_value = min(row_values)

                    # Apply highlighting
                    for col in range(2, ws.max_column + 1):
                        data_cell = ws.cell(row=row, column=col)
                        cell_value = data_cell.value

                        try:
                            val_str = str(cell_value).strip().replace('$', '').replace(',', '')
                            if '(' in val_str:
                                val_str = val_str.replace('(', '-').replace(')', '')
                            parsed_value = float(val_str)

                            # Highlight winner
                            if is_high_is_good and parsed_value == winner_value and winner_value != float('-inf'):
                                data_cell.fill = PatternFill(start_color=WINNER_GREEN, end_color=WINNER_GREEN, fill_type="solid")
                            elif is_low_is_good and parsed_value == winner_value and parsed_value != 0 and winner_value != float('inf'):
                                data_cell.fill = PatternFill(start_color=WINNER_GREEN, end_color=WINNER_GREEN, fill_type="solid")
                        except:
                            pass

                        data_cell.font = DATA_FONT
                        data_cell.alignment = Alignment(horizontal='center', vertical='center')
                        data_cell.border = THIN_BORDER
            else:
                # Non-colorable rows (like Rank) - just format, no highlighting
                for col in range(2, ws.max_column + 1):
                    data_cell = ws.cell(row=row, column=col)
                    data_cell.font = DATA_FONT
                    data_cell.alignment = Alignment(horizontal='center', vertical='center')
                    data_cell.border = THIN_BORDER

    freeze_top_rows(ws, num_rows=2)
    auto_width_columns(ws, min_width=15, max_width=25)
    add_color_legend(ws, 'value')


def format_cash_flow_sheet(workbook, sheet_name, period, year):
    """
    Format Cash Flow sheet with winner and above-average highlighting.
    - Green: Winner (highest cash flow = best)
    - Yellow: Above group average
    Matches web page logic from group_cash_flow.py.
    """
    ws = workbook[sheet_name]

    # Add metadata header
    add_metadata_header(ws, sheet_name, period, year)

    # Format column headers (now row 2)
    header_row = 2
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=header_row, column=col)
        cell.font = COLUMN_HEADER_FONT
        cell.fill = PatternFill(start_color=ATLAS_BLUE, end_color=ATLAS_BLUE, fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

    ws.row_dimensions[header_row].height = 20

    # Define items where higher is better (matching web logic)
    high_is_good_items = [
        'Operating Cash Flow (OCF)',
        'OCF/Revenue',
        'Net Cash Flow (NCF)',
        'NCF/Revenue',
        'Current Period Cash',
        'Actual Current Period Cash'
    ]

    # First pass - calculate group averages for colorable rows
    row_averages = {}
    for row in range(3, ws.max_row + 1):
        row_label = ws.cell(row=row, column=1).value
        if row_label:
            is_high_is_good = any(str(row_label).startswith(item) or str(row_label) == item for item in high_is_good_items)
            if is_high_is_good:
                row_values = []
                for col in range(2, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and cell_value != '-':
                        try:
                            val_str = str(cell_value).strip().replace('%', '').replace('$', '').replace(',', '')
                            if '(' in val_str:
                                val_str = val_str.replace('(', '-').replace(')', '')
                            row_values.append(float(val_str))
                        except:
                            pass
                if row_values:
                    row_averages[row] = sum(row_values) / len(row_values)

    # Format data rows with winner and above-average highlighting
    for row in range(3, ws.max_row + 1):
        row_label = ws.cell(row=row, column=1).value

        if row_label:
            # Bold the row label
            ws.cell(row=row, column=1).font = BOLD_DATA_FONT
            ws.cell(row=row, column=1).alignment = Alignment(horizontal='left', vertical='center')
            ws.cell(row=row, column=1).border = THIN_BORDER

            # Check if this row should be color-coded (higher is better)
            is_high_is_good = any(str(row_label).startswith(item) or str(row_label) == item for item in high_is_good_items)

            if is_high_is_good:
                # Find MAX value in the row (excluding column A)
                row_values = []
                for col in range(2, ws.max_column + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value and cell_value != '-':
                        try:
                            val_str = str(cell_value).strip().replace('%', '').replace('$', '').replace(',', '')
                            if '(' in val_str:
                                val_str = val_str.replace('(', '-').replace(')', '')
                            row_values.append(float(val_str))
                        except:
                            row_values.append(float('-inf'))
                    else:
                        row_values.append(float('-inf'))

                # Find winner (maximum value = best)
                max_value = max(row_values) if row_values else float('-inf')
                group_avg = row_averages.get(row, 0)

                for col_idx, col in enumerate(range(2, ws.max_column + 1)):
                    data_cell = ws.cell(row=row, column=col)
                    cell_value = data_cell.value

                    parsed_value = None
                    try:
                        val_str = str(cell_value).strip().replace('%', '').replace('$', '').replace(',', '')
                        if '(' in val_str:
                            val_str = val_str.replace('(', '-').replace(')', '')
                        parsed_value = float(val_str)
                    except:
                        pass

                    if parsed_value is not None and parsed_value != float('-inf'):
                        # Green for winner (highest)
                        if parsed_value == max_value and max_value != float('-inf'):
                            data_cell.fill = PatternFill(start_color=WINNER_GREEN, end_color=WINNER_GREEN, fill_type="solid")
                        # Yellow for above average
                        elif parsed_value > group_avg:
                            data_cell.fill = PatternFill(start_color=ABOVE_AVG_YELLOW, end_color=ABOVE_AVG_YELLOW, fill_type="solid")

                    data_cell.font = DATA_FONT
                    data_cell.alignment = Alignment(horizontal='center', vertical='center')
                    data_cell.border = THIN_BORDER
            else:
                # Non-colorable rows - just format, no highlighting
                for col in range(2, ws.max_column + 1):
                    data_cell = ws.cell(row=row, column=col)
                    data_cell.font = DATA_FONT
                    data_cell.alignment = Alignment(horizontal='center', vertical='center')
                    data_cell.border = THIN_BORDER

    freeze_top_rows(ws, num_rows=2)
    auto_width_columns(ws, min_width=15, max_width=25)
    add_color_legend(ws, 'cash_flow')
