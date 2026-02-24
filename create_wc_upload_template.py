"""
Script to create BPC W&C (Wins & Challenges) Upload Template Excel file
Creates an Excel file with Instructions, Wins, Challenges, and Action Items sheets
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter


def create_wc_excel_template():
    """Create the BPC W&C Upload Template Excel file"""

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create sheets in order
    create_instructions_sheet(wb)
    create_wins_sheet(wb)
    create_challenges_sheet(wb)
    create_action_items_sheet(wb)

    # Save the file
    output_path = 'bpc_upload_template/BPC2_WC_Upload_Template.xlsx'
    wb.save(output_path)
    print(f"Template created successfully: {output_path}")


def create_instructions_sheet(wb):
    """Create Instructions sheet with usage guide"""

    ws = wb.create_sheet("Instructions", 0)

    # Define styles
    title_font = Font(size=18, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="025a9a", end_color="025a9a", fill_type="solid")
    section_font = Font(size=12, bold=True, color="025a9a")
    body_font = Font(size=11)
    bullet_font = Font(size=11)
    tip_fill = PatternFill(start_color="e6f7ff", end_color="e6f7ff", fill_type="solid")

    # Row 1: Title
    ws['A1'] = "BPC2 Wins & Challenges Upload Template"
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:B1')
    ws.row_dimensions[1].height = 35

    # Row 3: Purpose section
    ws['A3'] = "Purpose"
    ws['A3'].font = section_font
    ws['A4'] = "Use this template to upload Wins, Challenges, and Action Items for a company."
    ws['A4'].font = body_font

    # Row 6: How to Use section
    ws['A6'] = "How to Use"
    ws['A6'].font = section_font

    instructions = [
        "1. Select your Company and Period in the dashboard BEFORE uploading",
        "2. Fill out the Wins, Challenges, and Action Items sheets (one item per row)",
        "3. Upload this file on the W&C Admin page",
        "4. Data will be saved as DRAFT (not visible on dashboard yet)",
        "5. Go to User Management > Publish Data to publish when ready",
    ]

    for i, instruction in enumerate(instructions):
        row = 7 + i
        ws[f'A{row}'] = instruction
        ws[f'A{row}'].font = bullet_font

    # Row 13: Sheet Descriptions section
    ws['A13'] = "Sheet Descriptions"
    ws['A13'].font = section_font

    sheet_descriptions = [
        ("Wins", "Positive achievements, successes, and strengths"),
        ("Challenges", "Areas for improvement, concerns, and obstacles"),
        ("Action Items", "Specific next steps and recommendations"),
    ]

    for i, (sheet_name, description) in enumerate(sheet_descriptions):
        row = 14 + i
        ws[f'A{row}'] = f"  {sheet_name}"
        ws[f'A{row}'].font = Font(size=11, bold=True)
        ws[f'B{row}'] = description
        ws[f'B{row}'].font = body_font

    # Row 18: Column Guide section
    ws['A18'] = "Column Guide"
    ws['A18'].font = section_font

    ws['A19'] = "  Column A (Description)"
    ws['A19'].font = Font(size=11, bold=True)
    ws['B19'] = "Enter the text content (required)"
    ws['B19'].font = body_font

    ws['A20'] = "  Column B (Display Order)"
    ws['A20'].font = Font(size=11, bold=True)
    ws['B20'] = "Number that controls the order shown on dashboard (1, 2, 3...)"
    ws['B20'].font = body_font

    # Row 22: Tips section
    ws['A22'] = "Tips"
    ws['A22'].font = section_font

    tips = [
        "Leave rows blank if you have fewer items to enter",
        "Display order is optional - will auto-assign if left empty",
        "Re-uploading will OVERWRITE existing drafts for the same company/period",
        "Maximum 5000 characters per item",
    ]

    for i, tip in enumerate(tips):
        row = 23 + i
        ws[f'A{row}'] = f"  {tip}"
        ws[f'A{row}'].font = bullet_font
        ws[f'A{row}'].fill = tip_fill
        ws[f'B{row}'].fill = tip_fill

    # Set column widths
    ws.column_dimensions['A'].width = 55
    ws.column_dimensions['B'].width = 50

    # Protect the instructions sheet
    ws.protection.sheet = True
    ws.protection.enable()


def create_wins_sheet(wb):
    """Create Wins data entry sheet"""
    ws = wb.create_sheet("Wins", 1)
    _setup_data_sheet(ws, "Wins", "Win Description", "Enter your wins below - one per row")


def create_challenges_sheet(wb):
    """Create Challenges data entry sheet"""
    ws = wb.create_sheet("Challenges", 2)
    _setup_data_sheet(ws, "Challenges", "Challenge Description", "Enter your challenges below - one per row")


def create_action_items_sheet(wb):
    """Create Action Items data entry sheet"""
    ws = wb.create_sheet("Action Items", 3)
    _setup_data_sheet(ws, "Action Items", "Action Item Description", "Enter your action items below - one per row")


def _setup_data_sheet(ws, sheet_title, col_a_header, instruction_text):
    """Common setup for Wins, Challenges, and Action Items sheets"""

    # Define styles
    title_font = Font(size=14, bold=True, color="FFFFFF")
    title_fill = PatternFill(start_color="025a9a", end_color="025a9a", fill_type="solid")
    header_font = Font(size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0e9cd5", end_color="0e9cd5", fill_type="solid")
    instruction_font = Font(size=10, italic=True, color="4a5568")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Row 1: Sheet title
    ws['A1'] = sheet_title
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:B1')
    ws.row_dimensions[1].height = 30

    # Row 2: Column headers
    ws['A2'] = col_a_header
    ws['A2'].font = header_font
    ws['A2'].fill = header_fill
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].border = thin_border

    ws['B2'] = "Display Order"
    ws['B2'].font = header_font
    ws['B2'].fill = header_fill
    ws['B2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['B2'].border = thin_border

    ws.row_dimensions[2].height = 22

    # Row 3: Instruction row
    ws['A3'] = instruction_text
    ws['A3'].font = instruction_font
    ws['A3'].alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[3].height = 20

    # Rows 4-23: Empty data rows with borders and formatting
    for row_num in range(4, 24):
        # Column A: Text entry
        cell_a = ws.cell(row=row_num, column=1)
        cell_a.border = thin_border
        cell_a.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        cell_a.protection = Protection(locked=False)

        # Column B: Display order
        cell_b = ws.cell(row=row_num, column=2)
        cell_b.border = thin_border
        cell_b.alignment = Alignment(horizontal='center', vertical='center')
        cell_b.protection = Protection(locked=False)
        # Pre-fill display order
        cell_b.value = row_num - 3

        ws.row_dimensions[row_num].height = 45  # Taller rows for multi-line text

    # Set column widths
    ws.column_dimensions['A'].width = 80  # Wide for text content
    ws.column_dimensions['B'].width = 15  # Narrow for order number

    # Freeze header rows
    ws.freeze_panes = 'A4'

    # Protect header rows but allow data entry
    ws.protection.sheet = True
    ws.protection.enable()


if __name__ == "__main__":
    create_wc_excel_template()
