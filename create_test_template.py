"""
Create a test Excel template with sample data to verify upload functionality
"""

from openpyxl import load_workbook

# Load the template
wb = load_workbook('bpc_upload_template/BPC1_Upload_Template.xlsx')

# Add sample data to Income Statement sheet
is_sheet = wb['Income Statement']

# Sample revenue data (starting at row 3, column C)
sample_is_data = [
    150000,   # Intra State HHG
    75000,    # Local HHG
    250000,   # Inter State HHG
    100000,   # Office & Industrial
    50000,    # Warehouse
    25000,    # Warehouse Handling
    80000,    # International
    40000,    # Packing & Unpacking
    15000,    # Booking & Royalties
    30000,    # Special Products
    10000,    # Records Storage
    20000,    # Military DPM Contracts
    45000,    # Distribution
    12000,    # Hotel Deliveries
    8000,     # Other Revenue
    # Skip blank row
    # Direct Expenses
    200000,   # Direct Wages
    120000,   # Vehicle Operating Expense
    60000,    # Packing/Warehouse Supplies
    80000,    # OO Exp Intra State
    90000,    # OO Inter State
    40000,    # OO O&I
    15000,    # OO Packing
    10000,    # OO Other
    20000,    # Claims
    35000,    # Other Trans Exp
    45000,    # Depreciation
    25000,    # Lease Expense Rev Equip
    18000,    # Rent
    12000,    # Other Direct Expenses
    # Skip blank row
    # Operating Expenses
    15000,    # Advertising/Marketing
    5000,     # Bad Debts
    35000,    # Sales Commissions
    2000,     # Contributions
    8000,     # Computer Support
    3000,     # Dues & Subscriptions
    55000,    # PR Taxes & Benefits
    4000,     # Equipment Leases Office Equip
    18000,    # Workman's Comp Insurance
    22000,    # Insurance
    12000,    # Legal & Accounting
    6000,     # Office Expense
    8000,     # Other Admin
    15000,    # Pension/Profit Sharing/401k
    10000,    # Professional Fees
    7000,     # Repairs & Maintenance
    85000,    # Salaries Admin
    5000,     # Taxes & Licenses
    9000,     # Tel/Fax/Utilities/Internet
    6000,     # Travel & Entertainment
    4000,     # Vehicle Expense Admin
    # Skip blank row
    # Other Income/Expenses
    2000,     # Other Income
    -120000,  # CEO Comp (negative)
    -3000,    # Other Expense (negative)
    -8000,    # Interest Expense (negative)
]

# Fill in IS data
row_num = 3
data_idx = 0
for row in is_sheet.iter_rows(min_row=3, max_row=is_sheet.max_row):
    cell_a = row[0]  # Line Item column
    cell_c = row[2]  # Amount column (12/31/25)

    # Skip if this is a header row (all caps) or blank row
    if cell_a.value and isinstance(cell_a.value, str):
        if cell_a.value.strip().isupper() or cell_a.value.strip() == '':
            continue
    elif not cell_a.value:
        continue

    # Add sample data
    if data_idx < len(sample_is_data):
        cell_c.value = sample_is_data[data_idx]
        data_idx += 1

# Add sample data to Balance Sheet sheet
bs_sheet = wb['Balance Sheet']

sample_bs_data = [
    # Current Assets
    150000,   # Cash & Cash Equivalents
    200000,   # Trade Accounts Receivable
    50000,    # Receivables
    25000,    # Other Receivables
    15000,    # Prepaid Expenses
    10000,    # Related Company Receivables
    8000,     # Owner Receivables
    12000,    # Other Current Assets
    # Skip blank row
    # Fixed Assets
    500000,   # Gross Fixed Assets
    -200000,  # Accumulated Depreciation (negative)
    # Skip blank row
    # Other Assets
    15000,    # Inter Company Receivable
    20000,    # Other Assets
    # Skip blank row
    # Current Liabilities
    75000,    # Notes Payable/Bank
    10000,    # Notes Payable/Owners
    120000,   # Trade Accounts Payable
    35000,    # Accrued Expenses
    25000,    # Current Portion LTD
    5000,     # Inter Company Payable
    15000,    # Other Current Liabilities
    # Skip blank row
    # Long-term Liabilities
    50000,    # EID Loan
    150000,   # Long-term Debt
    30000,    # Notes Payable Owners (LT)
    10000,    # Inter Company Debt
    8000,     # Other LT Liabilities
    # Skip blank row
    # Equity
    277000,   # Owner's Equity (calculated to balance)
]

# Fill in BS data
row_num = 3
data_idx = 0
for row in bs_sheet.iter_rows(min_row=3, max_row=bs_sheet.max_row):
    cell_a = row[0]  # Line Item column
    cell_c = row[2]  # Amount column (12/31/25)

    # Skip if this is a header row (all caps) or blank row
    if cell_a.value and isinstance(cell_a.value, str):
        if cell_a.value.strip().isupper() or cell_a.value.strip() == '':
            continue
    elif not cell_a.value:
        continue

    # Add sample data
    if data_idx < len(sample_bs_data):
        cell_c.value = sample_bs_data[data_idx]
        data_idx += 1

# Save test template
wb.save('bpc_upload_template/BPC1_Upload_Template_TEST.xlsx')
print("✅ Test template created: bpc_upload_template/BPC1_Upload_Template_TEST.xlsx")
print(f"   - Added {len(sample_is_data)} Income Statement values")
print(f"   - Added {len(sample_bs_data)} Balance Sheet values")
print("   - Ready for upload testing!")
