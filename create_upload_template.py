"""
Script to create BPC Upload Template Excel file
Creates an Excel file with Balance Sheet and Income Statement sheets
Each sheet has: Line Item, Description, and Date (12/31/25) columns
"""

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# Import descriptions
from description_mappings import BALANCE_SHEET_DESCRIPTIONS, INCOME_STATEMENT_DESCRIPTIONS


def create_excel_template():
    """Create the BPC Upload Template Excel file"""

    # Create workbook
    wb = Workbook()

    # Remove default sheet
    wb.remove(wb.active)

    # Create Instructions sheet first
    create_instructions_sheet(wb)

    # Create Income Statement sheet
    create_income_statement_sheet(wb)

    # Create Balance Sheet sheet
    create_balance_sheet_sheet(wb)

    # Save the file
    output_path = 'bpc_upload_template/BPC1_Upload_Template_NEW_With_Subtotals.xlsx'
    wb.save(output_path)
    print(f"✅ Template created successfully: {output_path}")


def create_instructions_sheet(wb):
    """Create Instructions sheet with user guidance"""

    ws = wb.create_sheet("Instructions", 0)

    atlas_blue = "025A9A"
    light_blue = "E8F4FD"
    dark_gray = "333333"

    title_font = Font(name="Montserrat", size=22, bold=True, color="FFFFFF")
    header_font = Font(name="Montserrat", size=16, bold=True, color=atlas_blue)
    number_font = Font(name="Montserrat", size=14, bold=True, color=dark_gray)
    body_font = Font(name="Montserrat", size=14, bold=True, color=dark_gray)
    note_bullet = Font(name="Montserrat", size=13, bold=True, color=dark_gray)
    note_font = Font(name="Montserrat", size=13, bold=True, italic=True, color="444444")

    title_fill = PatternFill(start_color=atlas_blue, end_color=atlas_blue, fill_type="solid")
    alt_fill = PatternFill(start_color=light_blue, end_color=light_blue, fill_type="solid")

    wrap_top = Alignment(wrap_text=True, vertical="top")

    # Column widths - A narrow for numbers, B-H wide for text
    ws.column_dimensions['A'].width = 5
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
        ws.column_dimensions[col_letter].width = 18

    # Row 1: Title
    ws.merge_cells('A1:H1')
    cell = ws['A1']
    cell.value = "BPC1 Upload Template \u2014 Instructions"
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 55

    # Row 2: Spacer
    ws.row_dimensions[2].height = 10

    # Row 3: Step 1 header
    ws.merge_cells('A3:H3')
    ws['A3'].value = "Step 1: Fill Out the Template"
    ws['A3'].font = header_font
    ws.row_dimensions[3].height = 40

    # Step 1 instructions
    step1_items = [
        ("1.", 'Open the "Income Statement" and "Balance Sheet" tabs in this workbook.'),
        ("2.", "Enter the dollar amounts for each line item in the Amount column (Column C). Each line item matches what you see on the dashboard's Income Statement and Balance Sheet pages."),
        ("3.", 'Enter raw numbers only \u2014 no dollar signs ($), commas, or other formatting. For example, enter 150000 not $150,000.'),
        ("4.", "Enter negative values with a minus sign (e.g., -5000 for accumulated depreciation)."),
        ("5.", "Enter 0 for any line item that does not apply to your company. Do not leave cells blank."),
        ("6.", "Do NOT modify the line item names in Column A or the descriptions in Column B. The system uses these names to match your data."),
        ("7.", "Subtotal rows (e.g., Total Revenue, Gross Profit, Total Assets) are calculated for you automatically as you enter values. You cannot edit these rows — they contain live formulas so you can sanity-check your numbers before uploading."),
    ]

    row = 4
    for num, text in step1_items:
        ws[f'A{row}'].value = num
        ws[f'A{row}'].font = number_font
        ws[f'A{row}'].alignment = Alignment(vertical="top", horizontal="right")
        ws.merge_cells(f'B{row}:H{row}')
        ws[f'B{row}'].value = text
        ws[f'B{row}'].font = body_font
        ws[f'B{row}'].alignment = wrap_top
        ws.row_dimensions[row].height = 58 if len(text) > 150 else 42
        row += 1

    # Spacer
    ws.row_dimensions[row].height = 10
    row += 1

    # Step 2 header
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].value = "Step 2: Upload Your Data"
    ws[f'A{row}'].font = header_font
    ws.row_dimensions[row].height = 40
    row += 1

    step2_items = [
        ("1.", "Log in to the BPC1 Dashboard and navigate to the Upload page."),
        ("2.", "At the top of the page, select your company name from the dropdown."),
        ("3.", "Select the correct reporting period \u2014 choose whether this is a Full Year or Mid Year report, and select the correct year."),
        ("4.", "On the sidebar of the upload page, you will find an option to upload this Excel file. Click it and select this file."),
        ("5.", "The system will parse your data and show you a preview. Review it carefully before submitting."),
        ("6.", 'Click the Submit button to upload your data. Once submitted, your data will be in "Submitted" status until an admin publishes it to the dashboard.'),
    ]

    for num, text in step2_items:
        ws[f'A{row}'].value = num
        ws[f'A{row}'].font = number_font
        ws[f'A{row}'].alignment = Alignment(vertical="top", horizontal="right")
        ws.merge_cells(f'B{row}:H{row}')
        ws[f'B{row}'].value = text
        ws[f'B{row}'].font = body_font
        ws[f'B{row}'].alignment = wrap_top
        ws.row_dimensions[row].height = 42
        row += 1

    # Spacer
    ws.row_dimensions[row].height = 10
    row += 1

    # Important Notes header
    ws.merge_cells(f'A{row}:H{row}')
    ws[f'A{row}'].value = "Important Notes"
    ws[f'A{row}'].font = header_font
    ws.row_dimensions[row].height = 40
    row += 1

    notes = [
        ("\u2022", "Make sure you select the correct company and period before uploading \u2014 the system will save data under whichever company and period you have selected."),
        ("\u2022", "This Instructions sheet is ignored during upload. Only the Income Statement and Balance Sheet sheets are processed."),
        ("\u2022", "If you see warnings after upload about unmatched line items, check that the line item names in your file have not been modified."),
    ]

    for bullet, text in notes:
        ws[f'A{row}'].value = bullet
        ws[f'A{row}'].font = note_bullet
        ws[f'A{row}'].alignment = Alignment(vertical="top", horizontal="center")
        ws.merge_cells(f'B{row}:H{row}')
        ws[f'B{row}'].value = text
        ws[f'B{row}'].font = note_font
        ws[f'B{row}'].alignment = wrap_top
        ws.row_dimensions[row].height = 45
        row += 1

    # Alternating row shading for readability
    step1_rows = list(range(4, 11))
    step2_rows = list(range(13, 19))
    note_rows = list(range(21, 24))
    all_content_rows = step1_rows + step2_rows + note_rows

    for i, r in enumerate(all_content_rows):
        if i % 2 == 0:
            for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
                ws[f'{col_letter}{r}'].fill = alt_fill

    # Protect sheet so users don't accidentally edit instructions
    ws.protection.sheet = True
    ws.protection.enable()
    ws.sheet_properties.tabColor = atlas_blue


def _h(label):
    return {'kind': 'header', 'label': label}


def _blank():
    return {'kind': 'blank'}


def _d(key, label, group=None):
    row = {'kind': 'data', 'key': key, 'label': label}
    if group:
        row['group'] = group
    return row


def _sum(name, label, group):
    return {'kind': 'subtotal_sum', 'name': name, 'label': label, 'group': group}


def _f(name, label, formula):
    return {'kind': 'subtotal_formula', 'name': name, 'label': label, 'formula': formula}


def create_income_statement_sheet(wb):
    """Create Income Statement sheet with all line items and live subtotal formulas."""

    ws = wb.create_sheet("Income Statement", 1)

    is_line_items = [
        _h('REVENUE'),
        _d('intra_state_hhg', 'Intra State HHG', 'rev'),
        _d('local_hhg', 'Local HHG', 'rev'),
        _d('inter_state_hhg', 'Inter State HHG', 'rev'),
        _d('office_industrial', 'Office & Industrial', 'rev'),
        _d('warehouse', 'Warehouse (Non-commercial)', 'rev'),
        _d('warehouse_handling', 'Warehouse Handling (Non-commercial)', 'rev'),
        _d('international', 'International', 'rev'),
        _d('packing_unpacking', 'Packing & Unpacking', 'rev'),
        _d('booking_royalties', 'Booking & Royalties', 'rev'),
        _d('special_products', 'Special Products', 'rev'),
        _d('records_storage', 'Records Storage', 'rev'),
        _d('military_dpm_contracts', 'Military DPM Contracts', 'rev'),
        _d('distribution', 'Distribution', 'rev'),
        _d('hotel_deliveries', 'Hotel Deliveries', 'rev'),
        _d('other_revenue', 'Other Revenue', 'rev'),
        _sum('TotalRev', 'Total Revenue', 'rev'),

        _blank(),
        _h('DIRECT EXPENSES (COST OF REVENUE)'),
        _d('direct_wages', 'Direct Wages', 'cor'),
        _d('vehicle_operating_expenses', 'Vehicle Operating Expense', 'cor'),
        _d('packing_warehouse_supplies', 'Packing/Warehouse Supplies', 'cor'),
        _d('oo_exp_intra_state', 'OO Exp Intra State', 'cor'),
        _d('oo_inter_state', 'OO Inter State', 'cor'),
        _d('oo_oi', 'OO O&I', 'cor'),
        _d('oo_packing', 'OO Packing', 'cor'),
        _d('oo_other', 'OO Other', 'cor'),
        _d('claims', 'Claims', 'cor'),
        _d('other_trans_exp', 'Other Trans Exp', 'cor'),
        _d('depreciation', 'Depreciation', 'cor'),
        _d('lease_expense_rev_equip', 'Lease Expense Rev Equip', 'cor'),
        _d('rent', 'Rent', 'cor'),
        _d('other_direct_expenses', 'Other Direct Expenses', 'cor'),
        _sum('TotalCOR', 'Total Cost of Revenue', 'cor'),
        _f('GrossProfit', 'Gross Profit', '=C{TotalRev}-C{TotalCOR}'),

        _blank(),
        _h('OPERATING EXPENSES'),
        _d('advertising_marketing', 'Advertising/Marketing', 'opex'),
        _d('bad_debts', 'Bad Debts', 'opex'),
        _d('sales_commissions', 'Sales Commissions', 'opex'),
        _d('contributions', 'Contributions', 'opex'),
        _d('computer_support', 'Computer Support', 'opex'),
        _d('dues_sub', 'Dues & Subscriptions', 'opex'),
        _d('pr_taxes_benefits', 'PR Taxes & Benefits', 'opex'),
        _d('equipment_leases_office_equip', 'Equipment Leases Office Equip', 'opex'),
        _d('workmans_comp_insurance', "Workman's Comp Insurance", 'opex'),
        _d('insurance', 'Insurance', 'opex'),
        _d('legal_accounting', 'Legal & Accounting', 'opex'),
        _d('office_expense', 'Office Expense', 'opex'),
        _d('other_admin', 'Other Admin', 'opex'),
        _d('pension_profit_sharing_401k', 'Pension/Profit Sharing/401k', 'opex'),
        _d('prof_fees', 'Professional Fees', 'opex'),
        _d('repairs_maint', 'Repairs & Maintenance', 'opex'),
        _d('salaries_admin', 'Salaries Admin', 'opex'),
        _d('taxes_licenses', 'Taxes & Licenses', 'opex'),
        _d('tel_fax_utilities_internet', 'Tel/Fax/Utilities/Internet', 'opex'),
        _d('travel_ent', 'Travel & Entertainment', 'opex'),
        _d('vehicle_expense_admin', 'Vehicle Expense Admin', 'opex'),
        _sum('TotalOpEx', 'Total Operating Expenses', 'opex'),
        _f('OpProfit', 'Operating Profit', '=C{GrossProfit}-C{TotalOpEx}'),

        _blank(),
        _h('OTHER INCOME/EXPENSES'),
        _d('other_income', 'Other Income', 'nonop'),
        _d('ceo_comp', 'CEO Comp', 'nonop'),
        _d('other_expense', 'Other Expense', 'nonop'),
        _d('interest_expense', 'Interest Expense', 'nonop'),
        _sum('TotalNonOp', 'Total Non-Operating', 'nonop'),
        _f('PBT', 'Profit Before Tax', '=C{OpProfit}+C{TotalNonOp}'),

        _blank(),
        _h('LABOR ANALYSIS'),
        _d('administrative_employees', 'Administrative Employees'),
        _d('number_of_branches', 'Number of Branches'),
    ]

    add_header_row(ws, "Income Statement Data Entry")
    add_data_rows(ws, is_line_items, INCOME_STATEMENT_DESCRIPTIONS)
    format_sheet(ws)


def create_balance_sheet_sheet(wb):
    """Create Balance Sheet sheet with all line items and live subtotal formulas."""

    ws = wb.create_sheet("Balance Sheet", 2)

    bs_line_items = [
        _h('CURRENT ASSETS'),
        _d('cash_and_cash_equivalents', 'Cash & Cash Equivalents', 'ca'),
        _d('trade_accounts_receivable', 'Trade Accounts Receivable', 'ca'),
        _d('receivables', 'Receivables', 'ca'),
        _d('other_receivables', 'Other Receivables', 'ca'),
        _d('prepaid_expenses', 'Prepaid Expenses', 'ca'),
        _d('related_company_receivables', 'Related Company Receivables', 'ca'),
        _d('owner_receivables', 'Owner Receivables', 'ca'),
        _d('other_current_assets', 'Other Current Assets', 'ca'),
        _sum('TCA', 'Total Current Assets', 'ca'),

        _blank(),
        _h('FIXED ASSETS'),
        _d('gross_fixed_assets', 'Gross Fixed Assets'),
        _d('accumulated_depreciation', 'Accumulated Depreciation'),
        _f('NFA', 'Net Fixed Assets', '=C{GrossFA}+C{AccumDep}'),

        _blank(),
        _h('OTHER ASSETS'),
        _d('inter_company_receivable', 'Inter Company Receivable', 'oa'),
        _d('other_assets', 'Other Assets', 'oa'),
        _sum('TOA', 'Total Other Assets', 'oa'),
        _f('TA', 'Total Assets', '=C{TCA}+C{NFA}+C{TOA}'),

        _blank(),
        _h('CURRENT LIABILITIES'),
        _d('notes_payable_bank', 'Notes Payable/Bank', 'cl'),
        _d('notes_payable_owners', 'Notes Payable/Owners', 'cl'),
        _d('trade_accounts_payable', 'Trade Accounts Payable', 'cl'),
        _d('accrued_expenses', 'Accrued Expenses', 'cl'),
        _d('current_portion_ltd', 'Current Portion LTD', 'cl'),
        _d('inter_company_payable', 'Inter Company Payable', 'cl'),
        _d('other_current_liabilities', 'Other Current Liabilities', 'cl'),
        _sum('TCL', 'Total Current Liabilities', 'cl'),

        _blank(),
        _h('LONG-TERM LIABILITIES'),
        _d('eid_loan', 'EID Loan', 'ltl'),
        _d('long_term_debt', 'Long-term Debt', 'ltl'),
        _d('notes_payable_owners_lt', 'Notes Payable Owners (LT)', 'ltl'),
        _d('inter_company_debt', 'Inter Company Debt', 'ltl'),
        _d('other_lt_liabilities', 'Other LT Liabilities', 'ltl'),
        _sum('TLTL', 'Total Long-term Liabilities', 'ltl'),

        _blank(),
        _h('EQUITY'),
        _d('owners_equity', "Owner's Equity"),

        _blank(),
        _f('TLE', 'Total Liabilities & Equity', '=C{TCL}+C{TLTL}+C{OwnersEquity}'),
        _f('BalanceCheck', 'Total Balance Check (Assets - Liab & Equity)', '=C{TA}-C{TLE}'),
    ]

    add_header_row(ws, "Balance Sheet Data Entry")
    # Balance sheet needs extra anchors for items not in a SUM group
    extra_anchors = {
        'gross_fixed_assets': 'GrossFA',
        'accumulated_depreciation': 'AccumDep',
        'owners_equity': 'OwnersEquity',
    }
    add_data_rows(ws, bs_line_items, BALANCE_SHEET_DESCRIPTIONS, extra_anchors=extra_anchors)
    format_sheet(ws)


def add_header_row(ws, sheet_title):
    """Add the header row with column names"""

    # Row 1: Sheet title
    ws['A1'] = sheet_title
    ws['A1'].font = Font(name="Montserrat", size=16, bold=True, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="025a9a", end_color="025a9a", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.merge_cells('A1:C1')
    ws.row_dimensions[1].height = 35

    # Row 2: Column headers
    headers = ['Line Item', 'Description', 'Amount']
    header_fill = PatternFill(start_color="0e9cd5", end_color="0e9cd5", fill_type="solid")
    header_font = Font(name="Montserrat", bold=True, color="FFFFFF", size=13)

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col_num)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

    ws.row_dimensions[2].height = 28


def add_data_rows(ws, line_items, description_dict, extra_anchors=None):
    """Add data rows with line items, descriptions, and amount column.

    Two-pass:
      Pass 1: assign row numbers, collect group first/last rows and subtotal anchors.
      Pass 2: write cells, resolving subtotal formulas against the anchor map.
    """

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    subtotal_fill = PatternFill(start_color="E8F4FD", end_color="E8F4FD", fill_type="solid")
    subtotal_font = Font(name="Montserrat", size=12, bold=True, color="025A9A")
    subtotal_top_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='medium', color="025A9A"),
        bottom=Side(style='thin')
    )

    number_validation = DataValidation(type="decimal", allow_blank=True)
    number_validation.error = "Please enter a number only. No dollar signs, commas, or text."
    number_validation.errorTitle = "Invalid Entry"
    number_validation.prompt = "Enter the dollar amount as a number (e.g., 150000 or -5000)"
    number_validation.promptTitle = "Amount"
    number_validation.showErrorMessage = True
    number_validation.showInputMessage = True
    ws.add_data_validation(number_validation)

    # ---- Pass 1: assign row numbers and collect anchors ----
    start_row = 3
    group_first = {}
    group_last = {}
    anchors = {}
    extra_anchors = extra_anchors or {}

    for i, item in enumerate(line_items):
        r = start_row + i
        item['_row'] = r
        kind = item['kind']
        if kind == 'data':
            grp = item.get('group')
            if grp:
                group_first.setdefault(grp, r)
                group_last[grp] = r
            # Allow anchoring individual data cells (e.g., Gross Fixed Assets)
            key = item.get('key')
            if key in extra_anchors:
                anchors[extra_anchors[key]] = r
        elif kind in ('subtotal_sum', 'subtotal_formula'):
            anchors[item['name']] = r

    # ---- Pass 2: write cells ----
    for item in line_items:
        row_num = item['_row']
        kind = item['kind']

        if kind == 'header':
            cell_a = ws.cell(row=row_num, column=1)
            cell_a.value = item['label']
            cell_a.font = Font(name="Montserrat", bold=True, size=12, color="FFFFFF")
            cell_a.fill = PatternFill(start_color="4a5568", end_color="4a5568", fill_type="solid")
            cell_a.alignment = Alignment(horizontal='left', vertical='center')
            ws.merge_cells(f'A{row_num}:C{row_num}')
            ws.row_dimensions[row_num].height = 25

        elif kind == 'blank':
            ws.row_dimensions[row_num].height = 8

        elif kind == 'data':
            field_key = item['key']
            field_label = item['label']

            cell_a = ws.cell(row=row_num, column=1)
            cell_a.value = field_label
            cell_a.font = Font(name="Montserrat", size=12, bold=True, color="333333")
            cell_a.alignment = Alignment(horizontal='left', vertical='center')
            cell_a.border = thin_border

            description = description_dict.get(field_key, '')
            cell_b = ws.cell(row=row_num, column=2)
            cell_b.value = description
            cell_b.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
            cell_b.font = Font(name="Montserrat", size=11.5, color="4a5568")
            cell_b.border = thin_border

            cell_c = ws.cell(row=row_num, column=3)
            cell_c.value = None
            cell_c.font = Font(name="Montserrat", size=12)
            cell_c.alignment = Alignment(horizontal='right', vertical='center')
            cell_c.border = thin_border
            cell_c.number_format = '#,##0'
            cell_c.protection = Protection(locked=False)
            number_validation.add(cell_c)

            desc_len = len(description) if description else 0
            if desc_len > 140:
                ws.row_dimensions[row_num].height = 56
            elif desc_len > 70:
                ws.row_dimensions[row_num].height = 40
            else:
                ws.row_dimensions[row_num].height = 28

        elif kind in ('subtotal_sum', 'subtotal_formula'):
            label = item['label']

            cell_a = ws.cell(row=row_num, column=1)
            cell_a.value = label
            cell_a.font = subtotal_font
            cell_a.fill = subtotal_fill
            cell_a.alignment = Alignment(horizontal='left', vertical='center')
            cell_a.border = subtotal_top_border

            cell_b = ws.cell(row=row_num, column=2)
            cell_b.value = None
            cell_b.fill = subtotal_fill
            cell_b.border = subtotal_top_border

            cell_c = ws.cell(row=row_num, column=3)
            if kind == 'subtotal_sum':
                grp = item['group']
                cell_c.value = f"=SUM(C{group_first[grp]}:C{group_last[grp]})"
            else:
                cell_c.value = item['formula'].format(**anchors)
            cell_c.font = subtotal_font
            cell_c.fill = subtotal_fill
            cell_c.alignment = Alignment(horizontal='right', vertical='center')
            cell_c.border = subtotal_top_border
            cell_c.number_format = '#,##0;(#,##0);"-"'
            # Subtotal cell stays locked (default) so the formula can't be overwritten
            cell_c.protection = Protection(locked=True)

            ws.row_dimensions[row_num].height = 28


def format_sheet(ws):
    """Apply formatting to the entire sheet"""

    # Set column widths
    ws.column_dimensions['A'].width = 35  # Line Item
    ws.column_dimensions['B'].width = 75  # Description (wider for full text)
    ws.column_dimensions['C'].width = 22  # Amount (fits 9-digit numbers with commas)

    # Freeze top two rows (title and headers) AND first two columns (Line Item and Description)
    # This allows users to scroll horizontally while keeping the line item labels visible
    ws.freeze_panes = 'C3'

    # Columns A and B stay locked by default. Column C data cells are unlocked
    # inside add_data_rows (subtotal cells stay locked so formulas can't be edited).

    # Enable worksheet protection
    # This prevents editing of locked cells (columns A and B) but allows editing of unlocked cells (column C)
    # Note: No password is set, so users can unprotect if needed, but it prevents accidental edits
    ws.protection.sheet = True
    ws.protection.enable()


if __name__ == "__main__":
    create_excel_template()
